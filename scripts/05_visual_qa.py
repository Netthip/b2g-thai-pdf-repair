#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
05_visual_qa.py — Build the page-by-page verification table.

Output: output/audit_reports/B2G_Page_by_Page_Verification.xlsx

Status legend:
  ผ่าน           = ตรวจเต็มหน้าแล้วไม่พบความผิดปกติ
  ไม่ผ่าน        = พบข้อผิดพลาดที่ยืนยันแล้ว (ดู Issue ID)
  ว่าง           = หน้าว่าง (ไม่มีเนื้อหา)
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OFFSET = 6

BLANK = {2, 4}
ERRORS = {66: ("TH-001", "หัวเรื่อง 'การปฏิรูปกฎหมาย' เพี้ยน (สระหาย+ร→ฐ)")}

HEAD = ["หน้า PDF", "หน้าที่พิมพ์", "จำนวนข้อผิดพลาดเดิม", "จำนวนที่แก้แล้ว",
        "จำนวนคงเหลือ", "Issue ID", "สถานะรอบ 1", "หมายเหตุ"]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "PageByPage"
    ws.append(HEAD)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for pdf in range(1, 73):
        printed = pdf - OFFSET
        printed = printed if printed >= 1 else "-"
        if pdf in ERRORS:
            iid, note = ERRORS[pdf]
            row = [pdf, printed, 1, 0, 1, iid, "ไม่ผ่าน", note]
            fill = "F8CBAD"
        elif pdf in BLANK:
            row = [pdf, printed, 0, 0, 0, "", "ว่าง", "หน้าว่าง"]
            fill = "D9D9D9"
        else:
            row = [pdf, printed, 0, 0, 0, "", "ผ่าน", "ตรวจเต็มหน้า 150 DPI ไม่พบความผิดปกติ"]
            fill = "E2EFDA"
        ws.append(row)
        for c in ws[ws.max_row]:
            c.fill = PatternFill("solid", fgColor=fill)
            c.alignment = Alignment(wrap_text=True, vertical="center")

    widths = [10, 12, 16, 14, 13, 10, 12, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    out = os.path.join(ROOT, "output/audit_reports/B2G_Page_by_Page_Verification.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print("written:", out)


if __name__ == "__main__":
    main()
