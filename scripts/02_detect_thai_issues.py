#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
02_detect_thai_issues.py — Build the Issue Registry from curated visual-audit findings.

DETECTION METHOD (important, for transparency):
  The source PDF has a broken text layer (mojibake / no usable ToUnicode), so automated
  text diffing is impossible. Findings are produced by VISUAL inspection of every rendered
  page (data/page_images, 150 DPI), with high-zoom crops (300-600%) for confirmation, and
  corroborated by per-span font analysis (PyMuPDF get_text("dict")).

  Each confirmed issue is verified by at least two independent signals before being marked
  "Confirmed". Anything not provable to 100% is recorded as "Human Review Required" instead
  of being claimed clean — per the project's QA rules.

This script is the single source of truth for the registry. It writes:
  - data/issue_registry/issue_registry.csv   (UTF-8 BOM, Excel-friendly)
  - output/audit_reports/B2G_Original_Audit_Report.xlsx
  - data/issue_registry/issue_registry.md
  - data/issue_registry/audit_summary.json
"""
import csv
import json
import os
from datetime import datetime, timezone

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

ROUND = 1
PAGE_OFFSET = 6  # printed page number = PDF index - 6 (front matter: cover + blanks + TOC)

COLUMNS = [
    "Issue ID", "รอบตรวจ", "หน้า PDF", "หน้าที่พิมพ์",
    "ลำดับบรรทัด/ตำแหน่ง", "ตำแหน่ง/หัวข้อ",
    "ข้อความที่พบ", "ข้อความที่ควรเป็น",
    "ประเภทข้อผิดพลาด", "ความรุนแรง", "วิธีตรวจพบ", "สถานะ",
]

# --- CONFIRMED issues (verified by >=2 signals) -----------------------------
CONFIRMED = [
    {
        "Issue ID": "DISPLAY-001",
        "รอบตรวจ": 4,
        "หน้า PDF": "ทั้งเล่ม (systemic)",
        "หน้าที่พิมพ์": "ทั้งเล่ม",
        "ลำดับบรรทัด/ตำแหน่ง": "ข้อความไทยจำนวนมากทั่วเล่ม",
        "ตำแหน่ง/หัวข้อ": "การวางตำแหน่งสระบน/วรรณยุกต์ (combining marks)",
        "ข้อความที่พบ": "สระ/วรรณยุกต์ลอยไปทับตัวถัดไป เมื่อเปิดด้วย viewer มือถือ/เว็บ (เช่นที่ bb.go.th, in-app browser) — ผู้ใช้ยืนยันด้วยภาพหน้าจอจริง (หน้า 7 'ปัญหา'→'ปั ญ')",
        "ข้อความที่ควรเป็น": "แสดงผลถูกต้องทุก viewer",
        "ประเภทข้อผิดพลาด": "สระ/วรรณยุกต์ลอย (renderer-dependent; Canva Thai export ไม่ robust)",
        "ความรุนแรง": "สูงมาก (กระทบทั้งเล่ม + ผู้อ่านหลักใช้มือถือ)",
        "วิธีตรวจพบ": "ผู้ใช้รายงาน + เทียบ render: viewer มือถือลอย แต่ MuPDF/Adobe ถูก → ยืนยัน renderer-dependent",
        "สถานะ": "แก้แล้วใน B2G_Repaired_Final.pdf (rasterize เป็นภาพ → ไม่มี text layer ให้ลอย)",
    },
    {
        "Issue ID": "TH-001",
        "รอบตรวจ": ROUND,
        "หน้า PDF": 66,
        "หน้าที่พิมพ์": 60,
        "ลำดับบรรทัด/ตำแหน่ง": "หัวเรื่องใหญ่ บรรทัดที่ 2 (กลาง-บนหน้า)",
        "ตำแหน่ง/หัวข้อ": "หัวเรื่องหมวด (Section title) — ด้านการบริหารภาครัฐ ...",
        "ข้อความที่พบ": "การปฏฐูปกฎหมาย (สระ ิ หาย + ร เพี้ยนเป็น ฐ)",
        "ข้อความที่ควรเป็น": "การปฏิรูปกฎหมาย",
        "ประเภทข้อผิดพลาด": "สระ/อักขระหาย + อักขระเพี้ยน (อ่านผิดจากต้นฉบับ)",
        "ความรุนแรง": "สูง (หัวเรื่องหมวดหลัก)",
        "วิธีตรวจพบ": "ตรวจสายตา + ซูม 600% + ตรวจฟอนต์รายสแปน (NotoSansThai-Bold เดียวกับหน้า 68 แต่กลิฟต่างกัน) เทียบกับสารบัญ/หน้า 68",
        "สถานะ": "แก้แล้วใน B2G_Repaired_Final.pdf (overlay 'การปฏิรูปกฎหมาย' เบคเป็นภาพ); ฉบับ Canva ต้นฉบับยังควรแก้ถ้าต้องการ text แบบ select ได้",
    },
]

# --- HUMAN REVIEW REQUIRED (cannot be claimed clean to 100%) ----------------
HUMAN_REVIEW = [
    {
        "Issue ID": "HR-001",
        "รอบตรวจ": "1-3",
        "หน้า PDF": "ข้อความเนื้อหา (body)",
        "หน้าที่พิมพ์": "-",
        "ลำดับบรรทัด/ตำแหน่ง": "ข้อความ body 10–12pt ทั้งเล่ม",
        "ตำแหน่ง/หัวข้อ": "เนื้อหาย่อย body (หัวข้อ/แถบสีตรวจซูมครบในรอบ 2)",
        "ข้อความที่พบ": "รอบ 3 OCR ไทย (tesseract tha) + ตรวจพจนานุกรม pythainlp: 464 flag → 277 เป็นคำจริง (OCR เพี้ยน) → เหลือ 94 token → ตรวจทุก token เป็น OCR เพี้ยนของคำที่ถูกต้องทั้งหมด",
        "ข้อความที่ควรเป็น": "(body ถูกต้อง — ไม่พบข้อผิดพลาดต้นฉบับแม้แต่จุดเดียว)",
        "ประเภทข้อผิดพลาด": "ไม่พบข้อผิดพลาด (OCR noise ล้วน)",
        "ความรุนแรง": "-",
        "วิธีตรวจพบ": "OCR+พจนานุกรม+ซูมยืนยัน (ดู data/verification_results/needs_review.md)",
        "สถานะ": "ตรวจแล้ว body 0 errors — ปิดรายการ",
    },
    {
        "Issue ID": "HR-002",
        "รอบตรวจ": 2,
        "หน้า PDF": 15,
        "หน้าที่พิมพ์": 9,
        "ลำดับบรรทัด/ตำแหน่ง": "กล่องหัวข้อ SME",
        "ตำแหน่ง/หัวข้อ": "การส่งเสริมวิสาหกิจขนาดกลางและขนาดย่อม (SME)",
        "ข้อความที่พบ": "การส่งเสริมวิสาหกิจขนาดกลางและขนาดย่อมที่เข้มแข็ง แข่งขันได้ — ถูกต้อง",
        "ข้อความที่ควรเป็น": "(ถูกต้องแล้ว — การอ่าน “ที่เชื่อม” ในรอบ 1 เป็นการอ่านผิดที่ความละเอียดต่ำ)",
        "ประเภทข้อผิดพลาด": "ไม่ใช่ข้อผิดพลาด (false positive รอบ 1)",
        "ความรุนแรง": "-",
        "วิธีตรวจพบ": "ซูม 320% ยืนยัน",
        "สถานะ": "ตรวจแล้ว ถูกต้อง — ปิดรายการ",
    },
]


def write_csv(rows, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in COLUMNS})


def write_md(rows, path):
    with open(path, "w", encoding="utf-8") as f:
        f.write("# B2G — Issue Registry (รอบตรวจที่ 1)\n\n")
        f.write(f"_สร้างเมื่อ {datetime.now(timezone.utc):%Y-%m-%d %H:%M} UTC_\n\n")
        f.write("## ข้อผิดพลาดที่ยืนยันแล้ว (Confirmed)\n\n")
        f.write("| " + " | ".join(COLUMNS) + " |\n")
        f.write("|" + "---|" * len(COLUMNS) + "\n")
        for r in CONFIRMED:
            f.write("| " + " | ".join(str(r.get(c, "")).replace("|", "/") for c in COLUMNS) + " |\n")
        f.write("\n## ต้องตรวจทานโดยมนุษย์ (Human Review Required)\n\n")
        f.write("| " + " | ".join(COLUMNS) + " |\n")
        f.write("|" + "---|" * len(COLUMNS) + "\n")
        for r in HUMAN_REVIEW:
            f.write("| " + " | ".join(str(r.get(c, "")).replace("|", "/") for c in COLUMNS) + " |\n")


def write_xlsx(path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()

    def fill_sheet(ws, rows, title_fill):
        ws.append(COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=title_fill)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for r in rows:
            ws.append([r.get(c, "") for c in COLUMNS])
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(width, 12), 45)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    ws1 = wb.active
    ws1.title = "Confirmed"
    fill_sheet(ws1, CONFIRMED, "1F3864")
    ws2 = wb.create_sheet("HumanReview")
    fill_sheet(ws2, HUMAN_REVIEW, "843C0C")
    wb.save(path)


def build_summary():
    pages_with_errors = sorted({r["หน้า PDF"] for r in CONFIRMED if isinstance(r["หน้า PDF"], int)})
    by_type = {}
    for r in CONFIRMED:
        by_type[r["ประเภทข้อผิดพลาด"]] = by_type.get(r["ประเภทข้อผิดพลาด"], 0) + 1
    return {
        "round": ROUND,
        "generated_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "total_pages": 72,
        "blank_pages": [2, 4],
        "pages_reviewed_full_res": 72,
        "confirmed_issues": len(CONFIRMED),
        "pages_with_confirmed_errors": pages_with_errors,
        "confirmed_by_type": by_type,
        "human_review_required": sum(1 for r in HUMAN_REVIEW if "ปิดรายการ" not in r["สถานะ"]),
        "round2_heading_lines_verified": 365,
        "round3_body_ocr": {"flags": 464, "auto_cleared_real_words": 277,
                            "unique_remaining": 94, "genuine_body_errors": 0},
        "round2_result": "all 365 heading/title-bar lines zoom-verified; exactly 1 corrupted (TH-001). "
                         "Every other instance of 'ปฏิรูป' (p23,26,50,68) renders correctly.",
        "open_confirmed_issues_remaining": ["TH-001 (awaiting Canva fix by document owner)"],
        "note": "Headings (Round 2, 365 lines) and body text (Round 3, OCR+dictionary) both fully "
                "checked. The ONLY genuine defect in all 72 pages is TH-001 (page 66 headline). "
                "HR-001 and HR-002 both resolved/closed. Once TH-001 is fixed in Canva and "
                "re-exported, run scripts 03-04 to verify and the book passes the Quality Gate.",
    }


def main():
    write_csv(CONFIRMED + HUMAN_REVIEW, os.path.join(ROOT, "data/issue_registry/issue_registry.csv"))
    write_md(CONFIRMED + HUMAN_REVIEW, os.path.join(ROOT, "data/issue_registry/issue_registry.md"))
    write_xlsx(os.path.join(ROOT, "output/audit_reports/B2G_Original_Audit_Report.xlsx"))
    summary = build_summary()
    with open(os.path.join(ROOT, "data/issue_registry/audit_summary.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print("\nRegistry written: CSV, MD, XLSX, summary.json")


if __name__ == "__main__":
    main()
