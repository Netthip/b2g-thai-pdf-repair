#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
10_numbers_audit.py — Audit every monetary number and propose corrected formatting.

Rule (per reviewer): do NOT change the source values. For figures in ล้านบาท, reformat to
the standard #,##0.0000 (4 decimal places + thousands separators). Leave counts (คน/แห่ง/ราย/
ไร่/กม.), percentages, per-person baht, years, etc. unchanged.

Values are read from the PDF text layer (digits are reliable even though Thai is mojibake);
the unit/context for each number is matched from the OCR text (data/ocr, readable Thai).

Outputs:
  data/issue_registry/numbers_audit.csv / .xlsx / .md
"""
import csv
import os
import re

import fitz

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SRC = os.path.join(ROOT, "input", "original_pdf", "เล่ม พรบ ประชาชน.pdf")
OCR = os.path.join(ROOT, "data", "ocr")

NUM = re.compile(r'\d{1,3}(?:,\d{3})+(?:\.\d+)?|\d+\.\d+')


COUNT_UNITS = ["คน", "ราย", "แห่ง", "ไร่", "กม", "กิโล", "ตัว", "ครัวเรือน", "โครงการ",
               "หน่วยงาน", "ชุมชน", "เครื่อง", "คัน", "บริษัท", "จุด", "ชุด", "หลัง",
               "เล่ม", "ฉบับ", "อัตรา", "ครั้ง", "ผืน", "สาขา"]


def classify(after, page_is_lb):
    """Find the NEAREST unit signal after the number and classify. Robust to Thai
    combining-char ordering: money = 'บาท' preceded by 'ล' (ล้านบาท) but NOT 'พัน'
    (พันล้านบาท = billions). 'ล้านคน' etc = count. standalone 'บาท' = per-person."""
    import unicodedata
    a = unicodedata.normalize("NFC", after).replace(" ", "")
    if after.lstrip()[:1] == "%":
        return "ร้อยละ/สัดส่วน", False
    cands = []  # (position, label, is_lb)
    # baht occurrences -> million / billion / per-person
    ib = a.find("บาท")
    if ib >= 0:
        pre = a[max(0, ib - 6):ib]
        if "พัน" in a[max(0, ib - 9):ib]:
            cands.append((ib, "พันล้านบาท", False))
        elif "ล" in pre:
            cands.append((ib, "ล้านบาท", True))
        else:
            cands.append((ib, "บาท (ต่อหน่วย)", False))
    # ลบ. and OCR garbles for ล้านบาท
    for kw in ("ลบ", "ź", "au"):
        i = a.find(kw)
        if i >= 0:
            cands.append((i, "ล้านบาท", True))
    # ล้าน + count noun
    for kw in ("ล้านคน", "ล้านราย", "ล้านไร่", "ล้านครัวเรือน", "ล้านตัน", "ล้านหน่วย"):
        i = a.find(kw)
        if i >= 0:
            cands.append((i, "จำนวน (ล้านหน่วย)", False))
    for u in COUNT_UNITS:
        i = a.find(u)
        if i >= 0:
            cands.append((i, "จำนวน (นับ)", False))
    if cands:
        cands.sort(key=lambda c: c[0])
        return cands[0][1], cands[0][2]
    if page_is_lb:
        return "ล้านบาท", True
    return "ไม่ทราบหน่วย", False


def to4(numstr):
    v = float(numstr.replace(",", ""))
    return f"{v:,.4f}"


def point4_from_right(numstr):
    """Reviewer's rule: keep the original digits, place the decimal so 4 digits follow it.
    e.g. '69,078.7' -> digits '690787' -> '69.0787'."""
    d = numstr.replace(",", "").replace(".", "")
    if len(d) <= 4:
        return "0." + d.zfill(4)
    ip = d[:-4].lstrip("0") or "0"
    return f"{int(ip):,}.{d[-4:]}"


def main():
    doc = fitz.open(SRC)
    rows = []
    for i in range(doc.page_count):
        text = doc[i].get_text()
        ocr = ""
        p = os.path.join(OCR, f"page_{i+1:02d}.txt")
        if os.path.exists(p):
            ocr = open(p, encoding="utf-8").read()
        oc = ocr.replace(" ", "")
        # default-to-ล้านบาท ONLY for header tables (unit in column header); elsewhere rely on
        # the per-number trailing unit so we don't mislabel GDP/billions/counts.
        page_is_lb = ("หน่วย:ล้านบาท" in oc) or ("หน่วยล้านบาท" in oc) or ("หน่วย;ล้านบาท" in oc)
        for m in NUM.finditer(text):
            n = m.group()
            if "," not in n and "." not in n:
                continue
            dec = len(n.split(".")[1]) if "." in n else 0
            # find context in OCR (text after the number = its unit, usually)
            ctx = ""; after = ""
            src = ocr; idx = ocr.find(n)
            if idx < 0:
                src = ocr.replace(" ", ""); idx = src.find(n)
            if idx >= 0:
                ctx = src[max(0, idx - 8):idx + len(n) + 18].replace("\n", " ")
                after = src[idx + len(n):idx + len(n) + 16].replace("\n", " ")
            unit, is_lb = classify(after, page_is_lb)
            cand = point4_from_right(n) if is_lb else ""
            rows.append({
                "หน้า PDF": i + 1, "หน้าพิมพ์": (i + 1 - 6) if i + 1 > 6 else "-",
                "ตัวเลขเดิม": n, "ทศนิยมเดิม": dec, "หน่วย(ตรวจ)": unit,
                "ถ้าวางจุด4จากขวา": cand, "ยืนยัน(ผู้ตรวจ)": "",
                "บริบท(OCR)": ctx.strip(),
            })
    cols = ["หน้า PDF", "หน้าพิมพ์", "ตัวเลขเดิม", "ทศนิยมเดิม", "หน่วย(ตรวจ)",
            "ถ้าวางจุด4จากขวา", "ยืนยัน(ผู้ตรวจ)", "บริบท(OCR)"]

    base = os.path.join(ROOT, "data", "issue_registry")
    with open(os.path.join(base, "numbers_audit.csv"), "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols); w.writeheader(); w.writerows(rows)

    # xlsx
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "Numbers"
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F3864")
    for r in rows:
        ws.append([r[c] for c in cols])
        if r["หน่วย(ตรวจ)"] == "ล้านบาท":
            for c in ws[ws.max_row]:
                c.fill = PatternFill("solid", fgColor="FFF2CC")
    widths = [9, 9, 16, 10, 16, 18, 8, 50]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = wdt
    ws.freeze_panes = "A2"
    wb.save(os.path.join(ROOT, "output", "audit_reports", "B2G_Numbers_Audit.xlsx"))

    # summary
    lb = [r for r in rows if r["หน่วย(ตรวจ)"] == "ล้านบาท"]
    from collections import Counter
    bytype = Counter(r["หน่วย(ตรวจ)"] for r in rows)
    md = ["# ตารางตรวจตัวเลขทั้งเล่ม\n", f"- ตัวเลขทั้งหมด: **{len(rows)}**",
          f"- ตรวจพบหน่วยล้านบาท: **{len(lb)}** (แถวเหลือง)",
          "- คอลัมน์ 'ถ้าวางจุด4จากขวา' = ค่าตามกฎผู้ตรวจ (เลขเดิม + จุดให้เหลือ 4 หลักหลังจุด) — ใช้เป็นตัวเลือกให้ยืนยัน",
          "- แยกตามหน่วย: " + ", ".join(f"{k}={v}" for k, v in bytype.most_common()), "",
          "| หน้า | ตัวเลขเดิม | ทศ. | หน่วย | ถ้าวางจุด4จากขวา | บริบท |",
          "|---|---|---|---|---|---|"]
    for r in rows:
        md.append(f"| {r['หน้า PDF']} | {r['ตัวเลขเดิม']} | {r['ทศนิยมเดิม']} | {r['หน่วย(ตรวจ)']} | "
                  f"{r['ถ้าวางจุด4จากขวา']} | {r['บริบท(OCR)'][:40]} |")
    open(os.path.join(base, "numbers_audit.md"), "w", encoding="utf-8").write("\n".join(md))

    print(f"total={len(rows)} | ล้านบาท={len(lb)}")
    print("by unit:", dict(bytype))


if __name__ == "__main__":
    main()
