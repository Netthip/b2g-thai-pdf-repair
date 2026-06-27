#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
12_final_error_list.py — Final consolidated list of WRONG ROWS (Thai-language errors only).

Scope (per reviewer's final decision):
  * Numbers are LEFT AS-IS (original) — not corrected, not guessed.
  * Only Thai-language readability errors are in scope.

Outputs: output/final_delivery/B2G_Final_Error_List.{csv,xlsx,md}
"""
import csv
import os

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUT = os.path.join(ROOT, "output", "final_delivery")

COLS = ["ลำดับ", "รหัส", "ประเภท", "หน้า PDF", "หน้าพิมพ์", "ตำแหน่ง",
        "ข้อความที่พบ", "ข้อความที่ควรเป็น", "จำนวนคำผิด", "สถานะการแก้"]

ROWS = [
    {
        "ลำดับ": 1, "รหัส": "TH-001", "ประเภท": "สระหาย/อักขระเพี้ยน (อ่านผิด)",
        "หน้า PDF": "66", "หน้าพิมพ์": "60", "ตำแหน่ง": "หัวเรื่องหมวด บรรทัดที่ 2",
        "ข้อความที่พบ": "การปฏฐูปกฎหมาย (สระ ิ หาย + ร→ฐ)",
        "ข้อความที่ควรเป็น": "การปฏิรูปกฎหมาย",
        "จำนวนคำผิด": "1", "สถานะการแก้": "แก้แล้วใน B2G_Repaired_Final.pdf",
    },
    {
        "ลำดับ": 2, "รหัส": "DISPLAY-001", "ประเภท": "สระ/วรรณยุกต์ลอย (เฉพาะ viewer มือถือ/เว็บ)",
        "หน้า PDF": "ทั้งเล่ม", "หน้าพิมพ์": "ทั้งเล่ม", "ตำแหน่ง": "ข้อความไทยทั่วเล่ม",
        "ข้อความที่พบ": "สระ/วรรณยุกต์ลอยไปทับตัวถัดไป เมื่อเปิดบนมือถือ/เว็บ",
        "ข้อความที่ควรเป็น": "แสดงผลถูกต้องทุก viewer",
        "จำนวนคำผิด": "ทั้งเล่ม (renderer-dependent — นับรายคำเป๊ะไม่ได้)",
        "สถานะการแก้": "แก้แล้ว (แปลงทุกหน้าเป็นภาพ → ไม่มี text layer ให้ลอย)",
    },
    {
        "ลำดับ": 3, "รหัส": "DATA-001", "ประเภท": "รูปแบบทศนิยมตัวเลขล้านบาท (ข้อสังเกต)",
        "หน้า PDF": "หลายหน้า", "หน้าพิมพ์": "-", "ตำแหน่ง": "ตัวเลขงบประมาณ",
        "ข้อความที่พบ": "รูปแบบทศนิยมไม่เป็นมาตรฐาน 4 ตำแหน่ง",
        "ข้อความที่ควรเป็น": "(ผู้ตรวจตัดสินใจคงตามต้นฉบับ — ไม่แก้ ไม่เดาค่า)",
        "จำนวนคำผิด": "ไม่นับ (คงต้นฉบับ)",
        "สถานะการแก้": "ไม่แก้ — บันทึกเป็นข้อสังเกต (ดู B2G_Numbers_Audit.xlsx)",
    },
]


def main():
    os.makedirs(OUT, exist_ok=True)
    with open(os.path.join(OUT, "B2G_Final_Error_List.csv"), "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLS); w.writeheader(); w.writerows(ROWS)
    with open(os.path.join(OUT, "B2G_Final_Error_List.md"), "w", encoding="utf-8") as f:
        f.write("# B2G — รายการข้อผิดพลาดสุดท้าย (เฉพาะภาษาไทย)\n\n")
        f.write("ขอบเขต: แก้เฉพาะภาษาไทยให้อ่านถูก · ตัวเลขคงตามต้นฉบับ (ไม่แก้/ไม่เดา)\n\n")
        f.write("| " + " | ".join(COLS) + " |\n|" + "---|" * len(COLS) + "\n")
        for r in ROWS:
            f.write("| " + " | ".join(str(r[c]).replace("|", "/") for c in COLS) + " |\n")
        f.write("\n## สรุปจำนวน\n")
        f.write("- คำผิดภาษาไทยเชิงเนื้อหา (ยืนยัน, แก้แล้ว): **1 คำ** (TH-001)\n")
        f.write("- ปัญหาสระลอย (การแสดงผล): **ทั้งเล่ม** (แก้แล้วด้วยการแปลงเป็นภาพ)\n")
        f.write("- ตัวเลข: คงตามต้นฉบับ ไม่แก้ (ตามผู้ตรวจ)\n")
        f.write("- เนื้อหา body อื่น: ตรวจ OCR ทั้งเล่มแล้ว = 0 คำผิด\n")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook(); ws = wb.active; ws.title = "WrongRows"
        ws.append(COLS)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F3864")
            c.alignment = Alignment(wrap_text=True, vertical="top")
        for r in ROWS:
            ws.append([r[c] for c in COLS])
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        for i, wd in enumerate([7, 12, 30, 10, 10, 24, 34, 28, 30, 32], 1):
            ws.column_dimensions[chr(64 + i)].width = wd
        wb.save(os.path.join(OUT, "B2G_Final_Error_List.xlsx"))
    except PermissionError:
        print("!! xlsx locked — skipped")
    print("Final error list written. Thai content errors = 1 (TH-001); display = whole-book (fixed); numbers = kept.")


if __name__ == "__main__":
    main()
